# 将LR1table输出为xml文件
import os
from openpyxl import Workbook
from utils import lr1table_output_dir

class LR1TableXml():
    def __init__(self, action, goto, table):
        """
        初始化LR1TableXml对象。

        :param action: LR1分析表的action表
        :param goto: LR1分析表的goto表
        :param table: LR1分析表的parse_table
        """
        # 初始化action表
        self.action = action
        # 初始化goto表
        self.goto = goto
        # 对action表进行排序
        self.action = sorted(self.action)
        # 对goto表进行排序
        self.goto = sorted(self.goto)
        # 初始化parse_table
        self.table = table
        # 初始化输出路径
        self.path = os.path.join(lr1table_output_dir,"LR1Table.xlsx")
        # 调用init方法
        self.init()

    def init(self):
        """
        初始化方法，用于创建并填充Excel表格。

        :return: 无
        """
        # 创建一个新的Excel工作簿
        wb = Workbook()
        # 获取默认的工作表
        sheet = wb['Sheet']
        # 在第一行第一列写入"状态"
        sheet.cell(1, 1, '状态')

        # 遍历action表中的每个元素
        for i in range(len(self.action) + len(self.action) - 1):
            # 如果i小于action表的长度
            if i < len(self.action):
                # 在第一行的第i+2列写入action表中的元素
                sheet.cell(0 + 1, i + 1 + 1, self.action[i])
            # 如果i大于等于action表的长度
            else:
                # 在第一行的第i+2列写入goto表中的元素
                sheet.cell(0 + 1, i + 1 + 1, self.goto[i - len(self.action)])

        # 遍历parse_table中的每个元素
        for i in range(len(self.table)):
            # 在第i+2行的第一列写入状态编号
            sheet.cell(i + 1 + 1, 0 + 1, str(i))
            # 遍历action表中的每个元素
            for item in self.action:
                # 如果parse_table中的元素不为空
                if self.table[i][item] != ' ':
                    # 在第i+2行的第action表中元素的索引+2列写入parse_table中的元素
                    sheet.cell(i + 1 + 1, self.action.index(item) + 1 + 1, self.table[i][item])
            # 遍历goto表中的每个元素
            for item in self.goto:
                # 如果parse_table中的元素不为空
                if self.table[i][item]:
                    # 在第i+2行的第goto表中元素的索引+action表长度+2列写入parse_table中的元素
                    sheet.cell(i + 1 + 1, self.goto.index(item) + 1 + len(self.action) + 1, str(self.table[i][item]))

        # 在第10行的第一列写入状态编号
        sheet.cell(9 + 1, 0 + 1, str(9))
        # 保存Excel文件
        wb.save(self.path)
